#!/usr/bin/env python
# coding: utf-8
"""
CV Data Extraction Pipeline
Extracts structured data from PDF CVs using Docling and OpenAI API.

Can be run as a script or via cronjob. All paths are absolute to ensure
consistent behavior regardless of working directory.
"""

import os
import sys
import json
import hashlib
import logging
from pathlib import Path
from datetime import datetime

# === SETUP ABSOLUTE PATHS ===
# Get the directory where this script is located
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

# Change working directory to script location for consistency
os.chdir(SCRIPT_DIR)

# Add project root to path for imports
sys.path.insert(0, str(PROJECT_ROOT))

# === LOGGING SETUP ===
LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(exist_ok=True)

log_filename = LOG_DIR / f"extract_cv_{datetime.now().strftime('%Y%m%d')}.log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

logger.info("=" * 60)
logger.info("Starting CV extraction pipeline")
logger.info(f"Script directory: {SCRIPT_DIR}")
logger.info(f"Project root: {PROJECT_ROOT}")

# === LOAD ENVIRONMENT ===
from dotenv import load_dotenv

# Load .env from project root (absolute path)
env_path = PROJECT_ROOT / ".env"
if env_path.exists():
    load_dotenv(env_path)
    logger.info(f"Loaded environment from: {env_path}")
else:
    logger.warning(f".env file not found at: {env_path}")

# Verify API key is available
api_key = os.getenv('OPENAI_API_KEY')
if not api_key:
    logger.error("OPENAI_API_KEY not found in environment variables")
    sys.exit(1)

# === IMPORTS ===
import pandas as pd
from openai import OpenAI, beta
from tqdm import tqdm
from pydantic import BaseModel
from typing import List, Optional

from src import config

# Initialize OpenAI client
client = OpenAI(api_key=api_key)

from docling.document_converter import DocumentConverter

# === HASH CACHE CONFIGURATION ===
# Use absolute path for hash file (in notebooks directory)
HASHES_FILE = SCRIPT_DIR / ".hashes.txt"
logger.info(f"Hash cache file: {HASHES_FILE}")


def calculate_file_hash(file_path):
    """Generate a SHA-256 hash of the entire file contents (binary)."""
    hash_sha256 = hashlib.sha256()

    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            hash_sha256.update(chunk)

    return hash_sha256.hexdigest()


def load_existing_hashes():
    """Load previously processed hashes from file."""
    if not HASHES_FILE.exists():
        return set()

    with open(HASHES_FILE, "r") as f:
        return set(line.strip() for line in f if line.strip())


def save_hash(hash_value):
    """Save a new hash to the file."""
    with open(HASHES_FILE, "a") as f:
        f.write(hash_value + "\n")


def process_pdf(file_path, force=False):
    """Check if a PDF has already been processed, if not (or if forced), process and save the hash."""
    file_hash = calculate_file_hash(file_path)
    existing_hashes = load_existing_hashes()

    if not force and file_hash in existing_hashes:
        return False

    logger.info(f"Processing new file: {file_path}")
    save_hash(file_hash)
    return True


# === PYDANTIC MODELS ===
class Educacion(BaseModel):
    institucion: str
    titulo: str
    fecha_inicio: Optional[str]
    fecha_fin: Optional[str]
    detalles: Optional[List[str]]


class Experiencia(BaseModel):
    empresa: str
    ubicacion: Optional[str]
    puesto: str
    fecha_inicio: Optional[str]
    fecha_fin: Optional[str]
    responsabilidades: Optional[List[str]]


class Habilidad(BaseModel):
    nombre: str
    nivel: Optional[str]


class Idioma(BaseModel):
    idioma: str
    nivel: Optional[str]


class Curriculum(BaseModel):
    nombre_completo: str
    correo: str
    telefono: Optional[str]
    resumen: Optional[str]
    experiencia: List[Experiencia]
    educacion: Optional[List[Educacion]]
    habilidades: Optional[List[Habilidad]]
    idiomas: Optional[List[Idioma]]
    certificaciones: Optional[List[str]]
    referencias: Optional[List[str]]


# === CONFIGURATION ===
# Allow override via environment variable, with fallback to default
DEFAULT_ROOT_DIR = os.path.join(
    os.path.expanduser('~'),
    'Library', 'CloudStorage', 'OneDrive-ESPARTINAS.A',
    'DocumentacionEspartina', 'INNOVACION',
    'Desarrollos propios', 'Base Datos CV Capital Humano'
)
root_dir = os.getenv('CV_ROOT_DIR', DEFAULT_ROOT_DIR)
output_name = os.getenv('CV_OUTPUT_NAME', 'base_cv_capital_humano.xlsx')

logger.info(f"Root directory: {root_dir}")
logger.info(f"Output file: {output_name}")

# Validate root directory exists
if not os.path.exists(root_dir):
    logger.error(f"Root directory does not exist: {root_dir}")
    logger.error("Set CV_ROOT_DIR environment variable to override")
    sys.exit(1)

# === DISCOVER PDF FILES ===
file_list = []

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith('.pdf'):
            file_path = os.path.join(root, file)
            file_list.append(file_path)

file_list = tuple(file_list)
logger.info(f"Found {len(file_list)} PDF files")

if len(file_list) == 0:
    logger.warning("No PDF files found in root directory")
    sys.exit(0)

# === LOAD EXISTING EXCEL DATA ===
query = "Eres un prolijo y laborioso data entry del sector de recursos humanos. Tu tarea es extraer muy detalladamente toda la información relevante de los curriculum vitae recibidos en formato pdf y pasarla prolijamente a una tabla excel con el formato dado."

excel_path = os.path.join(root_dir, output_name)
expected_sheets = ['Candidatos', 'Experiencia', 'Educacion', 'Habilidades', 'Certificaciones']

# Check if Excel file exists
if not os.path.exists(excel_path):
    logger.error(f"Excel file not found: {excel_path}")
    logger.error("Create the Excel file with required sheets first")
    sys.exit(1)

try:
    all_sheets = pd.read_excel(excel_path, sheet_name=None)
    loaded_dfs = {}

    for sheet in expected_sheets:
        if sheet in all_sheets:
            loaded_dfs[sheet] = all_sheets[sheet]
        else:
            logger.warning(f'Sheet "{sheet}" is not present in the file.')

    candidatos_df = loaded_dfs.get('Candidatos')

    if candidatos_df is None or candidatos_df.empty:
        logger.error("Candidatos sheet is empty or missing")
        sys.exit(1)

    logger.info(f"Loaded Excel file with {len(candidatos_df)} existing candidates")

except Exception as e:
    logger.error(f"Error reading Excel file: {e}")
    sys.exit(1)

# === PROCESS PDFs ===
processed = []
id_cv = candidatos_df.candidato_id.max() + 1
excluded_extensions = ('.png', '.xlsx', '.jpeg', '.jpg', '.gif')

# Initialize lists for each DataFrame
candidatos_data = []
experiencia_data = []
educacion_data = []
habilidades_data = []
certificaciones_data = []

# Count files to process
files_to_process = 0
files_skipped = 0
files_errored = 0

for file_path in tqdm(file_list, desc="Processing CVs"):
    if file_path.lower().endswith(excluded_extensions):
        continue

    if not process_pdf(file_path=file_path, force=False):
        files_skipped += 1
        continue

    files_to_process += 1

    # Convert PDF to Markdown
    try:
        source = file_path
        converter = DocumentConverter()
        result = converter.convert(source)
        result_text = result.document.export_to_markdown()
    except Exception as e:
        logger.error(f'Docling error for {file_path}: {e}')
        files_errored += 1
        continue

    # Extract structured data via OpenAI
    try:
        response = beta.chat.completions.parse(
            model="gpt-4o-mini-2024-07-18",
            messages=[
                {
                    "role": "user",
                    "content": f"Los datos del archivo {file_path} están en formato markdown:\n{result_text}\n\nPregunta: {query}"
                }
            ],
            temperature=0,
            max_tokens=15000,
            response_format=Curriculum,
            top_p=1
        )

        json_content = response.choices[0].message.content
        data = json.loads(json_content)

        candidato_id = id_cv
        id_cv += 1
        nombre_completo = data['nombre_completo']
        logger.info(f'Extracted data for: {nombre_completo}')

        # Candidate general data
        candidatos_data.append({
            'candidato_id': candidato_id,
            'nombre_completo': nombre_completo,
            'correo': data['correo'],
            'telefono': data['telefono'],
            'resumen': data['resumen'],
            'file_path': file_path
        })

        # Experience
        for exp in data.get('experiencia', []):
            experiencia_data.append({
                'candidato_id': candidato_id,
                'nombre_completo': nombre_completo,
                'empresa': exp['empresa'],
                'ubicacion': exp['ubicacion'],
                'puesto': exp['puesto'],
                'fecha_inicio': exp['fecha_inicio'],
                'fecha_fin': exp['fecha_fin'],
                'responsabilidades': ", ".join(exp['responsabilidades']) if exp['responsabilidades'] else None
            })

        # Education
        for edu in data.get('educacion', []):
            educacion_data.append({
                'candidato_id': candidato_id,
                'nombre_completo': nombre_completo,
                'institucion': edu['institucion'],
                'titulo': edu['titulo'],
                'fecha_inicio': edu['fecha_inicio'],
                'fecha_fin': edu['fecha_fin'],
                'detalles': ", ".join(edu['detalles']) if edu['detalles'] else None
            })

        # Skills
        for hab in data.get('habilidades', []):
            habilidades_data.append({
                'candidato_id': candidato_id,
                'nombre_completo': nombre_completo,
                'nombre': hab['nombre'],
                'nivel': hab['nivel']
            })

        # Certifications
        if data.get('certificaciones'):
            for cert in data['certificaciones']:
                certificaciones_data.append({
                    'candidato_id': candidato_id,
                    'nombre_completo': nombre_completo,
                    'certificacion': cert
                })

        processed.append(file_path)

    except Exception as e:
        logger.error(f'OpenAI API error for {file_path}: {e}')
        files_errored += 1

# === SUMMARY ===
logger.info("-" * 40)
logger.info(f"Processing summary:")
logger.info(f"  - Total PDFs found: {len(file_list)}")
logger.info(f"  - Already processed (skipped): {files_skipped}")
logger.info(f"  - Newly processed: {len(processed)}")
logger.info(f"  - Errors: {files_errored}")

# === SAVE TO EXCEL ===
if not candidatos_data:
    logger.info("No new candidates to add. Excel file unchanged.")
else:
    logger.info(f"Adding {len(candidatos_data)} new candidates to Excel...")

    # Prepare DataFrames
    candidatos_df_actual = pd.DataFrame(candidatos_data)
    candidatos_df_actual['zona/area'] = candidatos_df_actual['file_path'].str.split('/', expand=True)[10]
    candidatos_df_actual = candidatos_df_actual[['candidato_id', 'nombre_completo', 'zona/area', 'correo', 'telefono', 'resumen', 'file_path']]

    # Experience
    if not experiencia_data:
        experiencia_df_actual = pd.DataFrame(columns=['candidato_id', 'nombre_completo', 'empresa', 'ubicacion', 'puesto', 'anio_inicio', 'fecha_inicio', 'fecha_fin', 'responsabilidades'])
    else:
        experiencia_df_actual = pd.DataFrame(experiencia_data)
        if 'fecha_inicio' in experiencia_df_actual.columns:
            experiencia_df_actual['anio_inicio'] = experiencia_df_actual['fecha_inicio'].str.extract(r'(\d{4})')
        if 'nombre_completo' in experiencia_df_actual.columns:
            experiencia_df_actual.nombre_completo = experiencia_df_actual.nombre_completo.str.title()
        experiencia_df_actual = experiencia_df_actual[['candidato_id', 'nombre_completo', 'empresa', 'ubicacion', 'puesto', 'anio_inicio', 'fecha_inicio', 'fecha_fin', 'responsabilidades']]

    # Education
    if not educacion_data:
        educacion_df_actual = pd.DataFrame(columns=['candidato_id', 'nombre_completo', 'institucion', 'titulo', 'anio_inicio', 'fecha_inicio', 'fecha_fin', 'detalles'])
    else:
        educacion_df_actual = pd.DataFrame(educacion_data)
        if 'nombre_completo' in educacion_df_actual.columns:
            educacion_df_actual.nombre_completo = educacion_df_actual.nombre_completo.str.title()
        if 'fecha_inicio' in educacion_df_actual.columns:
            educacion_df_actual['anio_inicio'] = educacion_df_actual['fecha_inicio'].str.extract(r'(\d{4})')
        educacion_df_actual = educacion_df_actual[['candidato_id', 'nombre_completo', 'institucion', 'titulo', 'anio_inicio', 'fecha_inicio', 'fecha_fin', 'detalles']]

    # Skills
    if not habilidades_data:
        habilidades_df_actual = pd.DataFrame(columns=['candidato_id', 'nombre_completo', 'nombre', 'nivel'])
    else:
        habilidades_df_actual = pd.DataFrame(habilidades_data)
        if 'nombre_completo' in habilidades_df_actual.columns:
            habilidades_df_actual.nombre_completo = habilidades_df_actual.nombre_completo.str.title()

    # Certifications
    if not certificaciones_data:
        certificaciones_df_actual = pd.DataFrame(columns=['candidato_id', 'nombre_completo', 'certificacion'])
    else:
        certificaciones_df_actual = pd.DataFrame(certificaciones_data)
        if 'nombre_completo' in certificaciones_df_actual.columns:
            certificaciones_df_actual.nombre_completo = certificaciones_df_actual.nombre_completo.str.title()

    # Write to Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            existing_data = pd.read_excel(excel_path, sheet_name=None)

            if not candidatos_df_actual.empty:
                start_row = len(existing_data['Candidatos']) + 1
                candidatos_df_actual.to_excel(writer, sheet_name='Candidatos', startrow=start_row, index=False, header=False)

            if not experiencia_df_actual.empty:
                start_row = len(existing_data['Experiencia']) + 1
                experiencia_df_actual.to_excel(writer, sheet_name='Experiencia', startrow=start_row, index=False, header=False)

            if not educacion_df_actual.empty:
                start_row = len(existing_data['Educacion']) + 1
                educacion_df_actual.to_excel(writer, sheet_name='Educacion', startrow=start_row, index=False, header=False)

            if not habilidades_df_actual.empty:
                start_row = len(existing_data['Habilidades']) + 1
                habilidades_df_actual.to_excel(writer, sheet_name='Habilidades', startrow=start_row, index=False, header=False)

            if not certificaciones_df_actual.empty:
                start_row = len(existing_data['Certificaciones']) + 1
                certificaciones_df_actual.to_excel(writer, sheet_name='Certificaciones', startrow=start_row, index=False, header=False)

        # Apply formatting
        workbook = load_workbook(excel_path)
        bold_font = Font(bold=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.freeze_panes = 'C2'

            for cell in sheet[1]:
                cell.font = bold_font

            for col_idx, col in enumerate(sheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        workbook.save(excel_path)
        logger.info(f"Excel file updated successfully: {excel_path}")

    except PermissionError:
        logger.error(f"Cannot write to Excel file - it may be open in another application: {excel_path}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error writing to Excel file: {e}")
        sys.exit(1)

logger.info("Pipeline completed successfully")
logger.info("=" * 60)
